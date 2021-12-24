from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
import datetime as dt
from pathlib import Path
import os

data_ex = {
    'inquiryid': '20211202001',
    'date': '2021-12-02',
    'author': 'EdwardHsu',
    'author_tel': '02-2796-1122 ext:223',
    'company': '寶晶能源',
    'comp_contact': '徐郁鈞',
    'contact_tel': '02-2345-6789 ext:223',
    'category': 'PV_module',
    'quotas': [{
        'item_name': 'PV-Module',
        'item_spec': '335W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '340W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '350W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '500W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '1000W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '1200W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '2000W',
    }, {
        'item_name': 'PV-Module',
        'item_spec': '3000W',
    }]
}
BASE_DIR = Path(__file__).resolve().parent.parent.parent


def to_excel(data):
    wb_filedir = os.path.join(BASE_DIR, 'quotation', 'files', 'inquiry_example_2.xlsx')
    wb = load_workbook(wb_filedir)
    ws = wb['詢價單']

    ws.protection.password = '50791838'
    ws.protection.sheet = True
    ws['B3'] = data['date']  # 詢價時間

    ws['B5'] = "{}\n({})".format(data['author'], data['author_tel'])  # 詢價人 = 詢價單建立人
    ws['B5'].alignment = Alignment(wrap_text=True)  # 換行
    ws['E4'] = data['company']  # 報價公司
    ws['E5'] = "{}\n({})".format(data['comp_contact'], data['contact_tel'])  # 報價人 = 報價單建立人
    ws['E5'].alignment = Alignment(wrap_text=True)  # 換行
    ws['E3'].protection = Protection(locked=False)  # 報價時間解鎖可填寫
    rh = ws.row_dimensions[5]  # get dimension for row 3
    rh.height = 35  # value in points, there is no "auto"

    # 設定幣別只能用選項
    crrn_code = ['新台幣', '美元', '歐元', '人民幣', '日圓']
    list_option = '"{}"'.format(','.join(crrn_code))
    data_val = DataValidation(type="list", formula1=list_option)
    ws.add_data_validation(data_val)

    data_cnt = len(data['quotas'])  # 確定詢價品項數量

    # 資料大於10筆時，將既有品項向下移
    if data_cnt > 10:
        row_shift = data_cnt - 10
        ws.unmerge_cells(start_row=19, start_column=1, end_row=19, end_column=8)
        ws.move_range('A19:H30', rows=row_shift, translate=False)
        ws.merge_cells('A{0}:H{0}'.format(19 + row_shift))
        row_h = ws.row_dimensions[7].height
        ws.row_dimensions[19 + row_shift].height = row_h

    # 詢價品項填入表格內
    maxl_c = 0
    for idx, row in enumerate(data['quotas'], start=0):
        rowcnt = 9 + idx

        # 複製上一列表格格式
        for colcnt in range(8):
            from_style = ws.cell(row=rowcnt - 1, column=colcnt + 1)._style
            ws.cell(row=rowcnt, column=colcnt + 1)._style = from_style
            ws.cell(row=rowcnt, column=colcnt + 1).number_format = ws.cell(row=rowcnt - 1,
                                                                           column=colcnt + 1).number_format
        # 填入項次
        ws.cell(row=rowcnt, column=1, value=idx + 1)

        # 填入資料
        ws.cell(row=rowcnt, column=2, value=row['item_sn'])
        ws.cell(row=rowcnt, column=3, value=row['item_name'])
        ws.cell(row=rowcnt, column=4, value=row['item_mfg'])

        # 紀錄欄寬，待調整成最適大小
        len_c = len(str(ws.cell(row=rowcnt, column=3).value))

        if len_c > maxl_c:
            maxl_c = len_c

        ws.cell(row=rowcnt, column=5, value=1)  # 填入數量1
        data_val.add(ws['F{}'.format(rowcnt)])  # 設定幣別為選項

        ws.cell(row=rowcnt, column=6).protection = Protection(locked=False)  # 解鎖可填寫區域
        print(ws.cell(row=rowcnt, column=6), ws.cell(row=rowcnt, column=5).protection)
        ws.cell(row=rowcnt, column=7).protection = Protection(locked=False)  # 解鎖可填寫區域
        print(ws.cell(row=rowcnt, column=7), ws.cell(row=rowcnt, column=5).protection)
        ws.cell(row=rowcnt, column=8).value = "=E{}*G{}".format(rowcnt, rowcnt)

    # 調整欄寬
    ws.column_dimensions['C'].width = maxl_c * 1.7

    # 設定新增品項區塊範圍
    new_cell_area = 'B21:H30'

    # bj6
    if data_cnt > 10:
        row_shift = data_cnt - 10
        new_cell_area = 'B{}:H{}'.format(21 + row_shift, 30 + row_shift)

    # 新增品項區塊
    for row in ws[new_cell_area]:
        for idx, col in enumerate(row, start=1):
            col.protection = Protection(locked=False)
            # 數量固定1
            if idx == 4:
                col.value = '=IF(C{}="","",1)'.format(col.row)
                col.protection = Protection(locked=True)
            # 設定匯率為選項
            elif idx == 5:
                data_val.add(col)
            # 總價加公式
            elif idx == 7:
                col.value = '=IF(G{0}="","",E{0}*G{0})'.format(col.row)
                col.protection = Protection(locked=True)
    # 擷取當天時間的年月
    [year, month] = dt.datetime.now().strftime('%Y-%m').split('-')

    # 建立年月資料夾
    file_path = os.path.join(BASE_DIR, 'static', 'files', 'inquiry', 'output', year, month)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    print(file_path)
    file_name = "{}_{}_詢價單{}_{}.xlsx".format(
        data['company'], data['category'], data['inquiryid'], dt.datetime.now().strftime("%Y%m%d%H%M%S"))
    wb.save(os.path.join(BASE_DIR, file_path, file_name))
    return file_name


test_data = {
    'item_sn': 'PVAESDDDDD',
    'item_name': 'ITEM_History',
    'data': [{
        'item_sn': 'AAA',
        'item_name': 'ABDC',
        'item_mfg': 'DDD',
        'item_spec': 'AXDF',
        'cnt': 1,
        'price': 10210,
        'crnt': 'TWD',
        'xchgrt': 1.000,
        'TWD_price': 10210,
        'qdate': '2021-12-01',
    }, {
        'item_sn': 'CCC',
        'item_name': 'SXS',
        'item_mfg': 'SSS',
        'item_spec': 'ASD',
        'cnt': 1,
        'price': 10,
        'crnt': 'USD',
        'xchgrt': 30.000,
        'TWD_price': 30.000,
        'qdate': '2021-12-02',
    }, ],
}


def item_to_excel(data):
    # 開啟新的excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'

    # 標題欄位
    col_names = ['料號', '品項名稱', '類別', '製造商', '品項規格', '數量', '價格(原幣別)',
                 '幣別', '匯率', '價格(台幣)', '報價日期']
    for idx, col in enumerate(col_names, start=1):
        ws.cell(row=1, column=idx, value=col)

    # 寫入資料
    if len(data['data']) != 0:  # 如有資料才寫入
        for rowcnt, row in enumerate(data['data'], start=2):
            for colcnt, key in enumerate(row, start=1):
                ws.cell(row=rowcnt, column=colcnt, value=row[key])

    # 擷取當天時間的年月
    [year, month] = dt.datetime.now().strftime('%Y-%m').split('-')

    # 建立年月資料夾
    file_path = os.path.join(BASE_DIR, 'static', 'files', 'itemhistory', 'output', year, month)
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # 建立檔案名稱
    file_name = "{}{}_歷史報價_{}.xlsx".format(data['item_sn'], data['item_name'], dt.datetime.now().strftime("%Y%m%d%H%M%S"))

    # 存檔
    wb.save(os.path.join(BASE_DIR, file_path, file_name))
    return file_name